"""Unit: export carries provenance metadata and renders 'not available' explicitly (US3, FR-010)."""

from __future__ import annotations

import csv
import io
from datetime import datetime, timezone

from openpyxl import load_workbook

from src.models.common import NOT_AVAILABLE_TOKEN
from src.models.export_record import ExportFormat
from src.models.query import QueryFilters
from src.services.exporters import build_export_record, to_csv, to_xlsx
from src.services.pricing import assemble_rows
from tests.conftest import _vm_rows

RETRIEVED = datetime(2026, 1, 1, 12, 0, 0, tzinfo=timezone.utc)


def _record(fmt):
    rows = assemble_rows(_vm_rows(), currency="USD", retrieved_at=RETRIEVED)
    filters = QueryFilters(serviceName="Virtual Machines", armRegionName="eastus")
    return build_export_record(rows, filters, RETRIEVED, fmt), rows


def test_csv_has_metadata_and_not_available_token():
    record, _ = _record(ExportFormat.csv)
    text = to_csv(record).decode("utf-8")
    assert "# Source" in text
    assert "Azure Retail Prices API" in text
    assert "# Retrieved at" in text
    # D2s_v5 has no reservation -> explicit token, not 0/blank.
    assert NOT_AVAILABLE_TOKEN in text

    reader = list(csv.reader(io.StringIO(text)))
    # The header row for the table appears after the metadata block.
    assert any(row and row[0] == "Service" for row in reader)


def test_xlsx_has_metadata_sheet_and_pricing_sheet():
    record, rows = _record(ExportFormat.xlsx)
    wb = load_workbook(io.BytesIO(to_xlsx(record)))
    assert "Metadata" in wb.sheetnames
    assert "Pricing" in wb.sheetnames

    meta_values = [c.value for row in wb["Metadata"].iter_rows() for c in row]
    assert "Source" in meta_values
    assert "Azure Retail Prices API" in meta_values

    pricing = wb["Pricing"]
    header = [c.value for c in next(pricing.iter_rows(max_row=1))]
    assert header[0] == "Service"
    # One header row + one row per comparison row.
    assert pricing.max_row == len(rows) + 1
