"""Integration: exported file matches the on-screen comparison exactly (US3, Principle IV, FR-010)."""

from __future__ import annotations

import csv
import io

from openpyxl import load_workbook


def _screen_rows(api_client):
    resp = api_client.get(
        "/api/pricing",
        params={"serviceName": "Virtual Machines", "armRegionName": "eastus", "currencyCode": "USD"},
    )
    assert resp.status_code == 200
    return resp.json()["rows"]


def test_csv_export_matches_screen(api_client):
    rows = _screen_rows(api_client)
    resp = api_client.get(
        "/api/export/csv",
        params={"serviceName": "Virtual Machines", "armRegionName": "eastus", "currencyCode": "USD"},
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("text/csv")

    reader = list(csv.reader(io.StringIO(resp.content.decode("utf-8"))))
    header_idx = next(i for i, r in enumerate(reader) if r and r[0] == "Service")
    data_rows = [r for r in reader[header_idx + 1 :] if r]
    exported_skus = {r[2] for r in data_rows}  # "SKU" column
    assert exported_skus == {r["armSkuName"] for r in rows}


def test_xlsx_export_matches_screen(api_client):
    rows = _screen_rows(api_client)
    resp = api_client.get(
        "/api/export/xlsx",
        params={"serviceName": "Virtual Machines", "armRegionName": "eastus", "currencyCode": "USD"},
    )
    assert resp.status_code == 200

    wb = load_workbook(io.BytesIO(resp.content))
    pricing = wb["Pricing"]
    body = list(pricing.iter_rows(min_row=2, values_only=True))
    exported_skus = {r[2] for r in body}
    assert exported_skus == {r["armSkuName"] for r in rows}


def test_export_rejects_missing_service_filter(api_client):
    resp = api_client.get("/api/export/csv", params={"armRegionName": "eastus"})
    assert resp.status_code == 400
